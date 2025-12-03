import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =====================================================
# 엑셀 스타일 적용 관련
# =====================================================

def apply_style_to_range(ws, start_row, end_row, start_col, end_col,
                         header_row=None, total_rows=None, total_cols=None):
    """ 특정 범위에 스타일 적용 """
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    total_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
    total_font = Font(bold=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 헤더
            if header_row and r == header_row:
                cell.fill = header_fill
                cell.font = header_font

            # 합계 행
            if total_rows and r in total_rows:
                cell.fill = total_fill
                cell.font = total_font

            # 합계 열
            if total_cols and c in total_cols:
                cell.fill = total_fill
                cell.font = total_font


def autosize_columns(ws):
    """ 글자 다 보이게 여유 있게 열 너비 조정 """
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        # 최소 폭 15, 넉넉한 여유 +3
        adjusted_width = max(max_len + 3, 15)
        ws.column_dimensions[col_letter].width = adjusted_width


# =====================================================
# 공통: 연도-월 요약 표(서울시/강남구용)
# =====================================================

def make_year_month_summary(df, date_col="계약년월"):
    """
    연도별 한 행, 1~12월 열 + 연간합계 + 월평균(12개월) + 월평균(거래월 기준)
    """
    s = df[date_col].astype(str)
    year = s.str[:4].astype(int)
    month = s.str[4:6].astype(int)

    temp = pd.DataFrame({"연": year, "월": month})
    temp["건수"] = 1

    table = temp.groupby(["연", "월"])["건수"].sum().unstack(fill_value=0)

    # 1~12월 모두 보장
    for m in range(1, 13):
        if m not in table.columns:
            table[m] = 0
    table = table[sorted(table.columns)]

    # 합계 & 월평균
    table["연간합계"] = table.sum(axis=1)
    active_months = (table.iloc[:, 0:12] > 0).sum(axis=1).replace(0, pd.NA)

    table["월평균(12개월)"] = (table["연간합계"] / 12).round(2)
    table["월평균(거래월)"] = (table["연간합계"] / active_months).round(2)

    # 표시용 라벨
    table.index = [f"{y}년" for y in table.index]
    month_labels = [f"{m}월" for m in range(1, 13)]
    table.columns = month_labels + ["연간합계", "월평균(12개월)", "월평균(거래월)"]

    return table.reset_index().rename(columns={"index": "년도", "연": "년도"})


# =====================================================
# 분석 1: 강남구 상업용 빌딩 거래량
# =====================================================

def 분석_강남구(df):
    df = df[df["시군구"].str.contains("강남구")].copy()

    # 동, 연/월 추출
    s = df["계약년월"].astype(str)
    df["동"] = df["시군구"].str.split().str[2]
    df["연"] = s.str[:4]
    df["월"] = s.str[4:6]
    df["년월코드"] = df["연"].astype(int) * 100 + df["월"].astype(int)

    # 동별-월별 피벗 (가로: 2024년 6월 형태)
    monthly = df.groupby(["동", "년월코드"]).size().reset_index(name="건수")
    pivot_m = monthly.pivot(index="동", columns="년월코드", values="건수").fillna(0).astype(int)

    def fmt_ym(code):
        code = int(code)
        y = code // 100
        m = code % 100
        return f"{y}년 {m}월"

    pivot_m.columns = [fmt_ym(c) for c in pivot_m.columns]
    pivot_m["합계"] = pivot_m.sum(axis=1)
    pivot_m.loc["합계"] = pivot_m.sum()
    df_monthly = pivot_m.reset_index()

    # 동별-년도별 피벗
    yearly = df.groupby(["동", "연"]).size().reset_index(name="건수")
    pivot_y = yearly.pivot(index="동", columns="연", values="건수").fillna(0).astype(int)
    pivot_y.columns = [f"{c}년" for c in pivot_y.columns]
    df_yearly = pivot_y.reset_index()

    # 강남 전체 연도-월 요약표 (노션 스샷 느낌)
    summary = make_year_month_summary(df, "계약년월")
    df_summary = summary  # 이미 reset_index 상태

    return df_monthly, df_yearly, df_summary


# =====================================================
# 분석 2: 금액대별 6개구
# =====================================================

def 분석_금액대(df):
    target_gu = ["강남구", "성동구", "종로구", "중구", "용산구", "마포구"]

    df["구"] = df["시군구"].str.extract(r"서울특별시 (\S+구)")
    df["금액"] = (
        df["거래금액(만원)"].astype(str).replace(",", "", regex=True).astype(float) / 10000
    )

    def 구간(x):
        if x < 50:
            return "50억 미만"
        elif x < 100:
            return "50~100억 미만"
        elif x < 200:
            return "100~200억 미만"
        elif x < 400:
            return "200~400억 미만"
        elif x < 1000:
            return "400억 이상"
        else:
            return "1000억 이상"

    df["금액구간"] = df["금액"].apply(구간)
    df = df[df["구"].isin(target_gu)]

    grouped = df.groupby(["구", "금액구간"]).size().reset_index(name="건수")
    pivot = grouped.pivot(index="구", columns="금액구간", values="건수").fillna(0).astype(int)

    # 열 순서 정렬
    col_order = [
        "50억 미만",
        "50~100억 미만",
        "100~200억 미만",
        "200~400억 미만",
        "400억 이상",
        "1000억 이상",
    ]
    pivot = pivot.reindex(columns=col_order)

    pivot["합계"] = pivot.sum(axis=1)
    pivot.loc["합계"] = pivot.sum()

    return pivot.reset_index()


# =====================================================
# 분석 3: 서울시 상업용 부동산 월별/년도별 거래량
# =====================================================

def 분석_서울시(df):
    # 서울 전체 (상업/업무 구분 없이 전체라고 하셨으니 필터 X)
    summary = make_year_month_summary(df, "계약년월")
    return summary


# =====================================================
# 헤더 자동 찾기
# =====================================================

def read_excel_with_auto_header(path):
    temp = pd.read_excel(path, header=None)
    for i in range(len(temp)):
        if "시군구" in temp.iloc[i].astype(str).values:
            return pd.read_excel(path, header=i)
    raise ValueError("헤더 찾기 실패")


# =====================================================
# 엑셀 여러 시트에 테이블 쓰기 + 스타일 정보 저장
# =====================================================

def write_table(writer, sheet_name, df, start_row, style_tasks,
                has_total_row=True, total_col_name="합계"):
    """
    writer: pd.ExcelWriter
    start_row: 1-based (엑셀 기준)
    style_tasks: {sheet: [(sr,er,sc,ec,header,total_rows,total_cols), ...]}
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row - 1)
    end_row = start_row + len(df)  # 헤더포함 len(df)+1 만큼 내려감
    ncols = len(df.columns)

    total_rows = [end_row] if has_total_row else []
    total_cols = []
    if total_col_name and total_col_name in df.columns:
        total_cols = [df.columns.get_loc(total_col_name) + 1]

    style_tasks.setdefault(sheet_name, []).append(
        (start_row, end_row, 1, ncols, start_row, total_rows, total_cols)
    )

    return end_row


# =====================================================
# GUI
# =====================================================

selected_files = []

def add_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    for f in files:
        if f not in selected_files:
            selected_files.append(f)
            listbox.insert(tk.END, f)

def remove_file():
    sel = listbox.curselection()
    if sel:
        idx = sel[0]
        listbox.delete(idx)
        selected_files.pop(idx)

def start_analysis():
    if not selected_files:
        messagebox.showerror("오류", "파일을 선택하세요.")
        return

    do1 = var1.get()
    do2 = var2.get()
    do3 = var3.get()

    if not (do1 or do2 or do3):
        messagebox.showerror("오류", "최소 1개의 분석을 선택하세요.")
        return

    combined = pd.DataFrame()
    try:
        for f in selected_files:
            combined = pd.concat([combined, read_excel_with_auto_header(f)], ignore_index=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_실거래가_분석결과.xlsx"

        style_tasks = {}

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # 1) 강남구 시트
            if do1:
                sheet = "강남구"
                cur_row = 1
                df_m, df_y, df_s = 분석_강남구(combined.copy())

                # 동별 월별
                end = write_table(writer, sheet, df_m, cur_row, style_tasks,
                                  has_total_row=True, total_col_name="합계")
                cur_row = end + 2

                # 동별 년도별
                end = write_table(writer, sheet, df_y, cur_row, style_tasks,
                                  has_total_row=False, total_col_name=None)
                cur_row = end + 2

                # 강남 전체 연도-월 요약
                end = write_table(writer, sheet, df_s, cur_row, style_tasks,
                                  has_total_row=False, total_col_name=None)

            # 2) 금액대별 시트
            if do2:
                sheet = "금액대별"
                cur_row = 1
                df_price = 분석_금액대(combined.copy())
                write_table(writer, sheet, df_price, cur_row, style_tasks,
                            has_total_row=True, total_col_name="합계")

            # 3) 서울시 시트
            if do3:
                sheet = "서울시"
                cur_row = 1
                df_seoul = 분석_서울시(combined.copy())
                write_table(writer, sheet, df_seoul, cur_row, style_tasks,
                            has_total_row=False, total_col_name=None)

        # 스타일 적용
        wb = load_workbook(filename)
        for sheet_name, ranges in style_tasks.items():
            ws = wb[sheet_name]
            for (sr, er, sc, ec, hr, total_rows, total_cols) in ranges:
                apply_style_to_range(ws, sr, er, sc, ec, hr, total_rows, total_cols)
            autosize_columns(ws)
        wb.save(filename)

        messagebox.showinfo("완료", f"분석이 완료되었습니다.\n\n저장 파일:\n{filename}")

    except Exception as e:
        messagebox.showerror("에러", str(e))


# =====================================================
# Tkinter GUI 생성
# =====================================================

root = tk.Tk()
root.title("빌딩 매각 현황 분석 프로그램")

# (간단 DPI 스케일 업 – 화질/글씨 작을 때 도움)
try:
    root.tk.call('tk', 'scaling', 1.3)
except Exception:
    pass

tk.Label(root, text="엑셀 파일을 선택하세요", font=("맑은 고딕", 10)).pack(pady=5)

listbox = tk.Listbox(root, width=80, height=8, font=("맑은 고딕", 9))
listbox.pack()

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Button(frame, text="파일 추가 (+)", command=add_files,
          font=("맑은 고딕", 9)).grid(row=0, column=0, padx=5)
tk.Button(frame, text="파일 삭제 (-)", command=remove_file,
          font=("맑은 고딕", 9)).grid(row=0, column=1, padx=5)

tk.Label(root, text="실행할 분석 선택:", font=("맑은 고딕", 10, "bold")).pack(pady=10)

var1 = tk.IntVar()
var2 = tk.IntVar()
var3 = tk.IntVar()

tk.Checkbutton(root, text="① 강남구 상업용 빌딩 거래량", variable=var1,
               font=("맑은 고딕", 9)).pack(anchor="w", padx=30)
tk.Checkbutton(root, text="② 금액대별 6개구 거래량", variable=var2,
               font=("맑은 고딕", 9)).pack(anchor="w", padx=30)
tk.Checkbutton(root, text="③ 서울시 상업용 부동산 월별/년도별 거래량", variable=var3,
               font=("맑은 고딕", 9)).pack(anchor="w", padx=30)

tk.Button(root, text="분석 시작", bg="#2e8b57", fg="black",
          font=("맑은 고딕", 10, "bold"), width=20,
          command=start_analysis).pack(pady=20)

root.mainloop()
